import os
import pandas as pd

from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook

ruta_completa = os.getcwd()

wb = load_workbook(ruta_completa +"/Registro_Imagenes.xlsx")
wb1 = wb['Hoja1']

def move2gray(val):
    print("El valor  que se movera a grayscale es "+val)
    row = 0
    lastrow_color = 1
    lastrow_grayscale = 1
    for i in range(3500,7000,1):
        name = wb1.cell(i,8).value
        subname = str(name)[-36:-4]
        if subname == val:
            row = i
    for i in range(2,3500,1):
            if wb1.cell(i,8).value != None:
                lastrow_grayscale = lastrow_grayscale + 1

    for i in range(3500,7000,1):
            if wb1.cell(i,8).value != None:
                lastrow_color = lastrow_color + 1
    lastrow_color = lastrow_color + 3500 - 1

    if row == 0:
        print("Resultado: El elemento buscado no se encuentra")
    elif row < 3500:
        print("Resultado: El elemento ya se encuentra en el dominio GRAYCOLOR")
    else:
        wb1.cell(lastrow_grayscale+1,2).value = wb1.cell(row,2).value
        wb1.cell(lastrow_grayscale+1,3).value = wb1.cell(row,3).value 
        wb1.cell(lastrow_grayscale+1,4).value = wb1.cell(row,4).value 
        wb1.cell(lastrow_grayscale+1,5).value = wb1.cell(row,5).value 
        wb1.cell(lastrow_grayscale+1,6).value = wb1.cell(row,6).value 
        wb1.cell(lastrow_grayscale+1,7).value = wb1.cell(row,7).value 
        wb1.cell(lastrow_grayscale+1,8).value = wb1.cell(row,8).value 
        wb1.cell(row,2).value = None
        wb1.cell(row,3).value = None
        wb1.cell(row,4).value = None
        wb1.cell(row,5).value = None
        wb1.cell(row,6).value = None
        wb1.cell(row,7).value = None
        wb1.cell(row,8).value = None
        for i in range(row,lastrow_color,1):
            wb1.cell(i,2).value = wb1.cell(i+1,2).value
            wb1.cell(i,3).value = wb1.cell(i+1,3).value 
            wb1.cell(i,4).value = wb1.cell(i+1,4).value 
            wb1.cell(i,5).value = wb1.cell(i+1,5).value 
            wb1.cell(i,6).value = wb1.cell(i+1,6).value 
            wb1.cell(i,7).value = wb1.cell(i+1,7).value 
            wb1.cell(i,8).value = wb1.cell(i+1,8).value 
        wb1.cell(lastrow_color,2).value = None
        wb1.cell(lastrow_color,3).value = None
        wb1.cell(lastrow_color,4).value = None
        wb1.cell(lastrow_color,5).value = None
        wb1.cell(lastrow_color,6).value = None
        wb1.cell(lastrow_color,7).value = None
        wb1.cell(lastrow_color,8).value = None
        wb.save(ruta_completa +"/Registro_Imagenes.xlsx")
        print("Resultado: Se ha movido a GRAYSCALE")
    print("--------------------")

def move2color(val):
    print("El valor que se movera a color es "+val)
    row = 0
    lastrow_color = 1
    lastrow_grayscale = 1
    for i in range(2,3500,1):
        name = wb1.cell(i,8).value
        subname = str(name)[-36:-4]
        if subname == val:
            row = i
    for i in range(2,3500,1):
            if wb1.cell(i,8).value != None:
                lastrow_grayscale = lastrow_grayscale + 1
    for i in range(3500,7000,1):
            if wb1.cell(i,8).value != None:
                lastrow_color = lastrow_color + 1
    lastrow_color = lastrow_color + 3500 - 1
    if row == 0:
        print("Resultado: El elemento buscado no se encuentra")
    elif row < 3500:
        wb1.cell(lastrow_color+1,2).value = wb1.cell(row,2).value
        wb1.cell(lastrow_color+1,3).value = wb1.cell(row,3).value 
        wb1.cell(lastrow_color+1,4).value = wb1.cell(row,4).value 
        wb1.cell(lastrow_color+1,5).value = wb1.cell(row,5).value 
        wb1.cell(lastrow_color+1,6).value = wb1.cell(row,6).value 
        wb1.cell(lastrow_color+1,7).value = wb1.cell(row,7).value 
        wb1.cell(lastrow_color+1,8).value = wb1.cell(row,8).value 
        wb1.cell(row,2).value = None
        wb1.cell(row,3).value = None
        wb1.cell(row,4).value = None
        wb1.cell(row,5).value = None
        wb1.cell(row,6).value = None
        wb1.cell(row,7).value = None
        wb1.cell(row,8).value = None
        for i in range(row,lastrow_grayscale,1):
            wb1.cell(i,2).value = wb1.cell(i+1,2).value
            wb1.cell(i,3).value = wb1.cell(i+1,3).value 
            wb1.cell(i,4).value = wb1.cell(i+1,4).value 
            wb1.cell(i,5).value = wb1.cell(i+1,5).value 
            wb1.cell(i,6).value = wb1.cell(i+1,6).value 
            wb1.cell(i,7).value = wb1.cell(i+1,7).value 
            wb1.cell(i,8).value = wb1.cell(i+1,8).value 
        wb1.cell(lastrow_grayscale,2).value = None
        wb1.cell(lastrow_grayscale,3).value = None
        wb1.cell(lastrow_grayscale,4).value = None
        wb1.cell(lastrow_grayscale,5).value = None
        wb1.cell(lastrow_grayscale,6).value = None
        wb1.cell(lastrow_grayscale,7).value = None
        wb1.cell(lastrow_grayscale,8).value = None
        wb.save(ruta_completa +"/Registro_Imagenes.xlsx")
        print("Resultado: Se ha movido a COLOR")
    else:
        print("Resultado: El elemento ya se encuentra en el dominio COLOR")
    print("--------------------")


def suppress(val):
    print("Se procede a eliminar el valor: "+val)
    row = 0
    lastrow = 1
    for i in range(2,7000,1):
        name = wb1.cell(i,8).value
        subname = str(name)[-36:-4]
        if subname == val:
            row = i
    if row == 0:
        print("Resultado: El elemento buscado no se encuentra")
    elif row < 3500:
        for i in range(2,3500,1):
            if wb1.cell(i,8).value != None:
                lastrow = lastrow + 1
        wb1.cell(row,2).value = None
        wb1.cell(row,3).value = None
        wb1.cell(row,4).value = None
        wb1.cell(row,5).value = None
        wb1.cell(row,6).value = None
        wb1.cell(row,7).value = None
        wb1.cell(row,8).value = None
        for i in range(row,lastrow,1):
            wb1.cell(i,2).value = wb1.cell(i+1,2).value
            wb1.cell(i,3).value = wb1.cell(i+1,3).value 
            wb1.cell(i,4).value = wb1.cell(i+1,4).value 
            wb1.cell(i,5).value = wb1.cell(i+1,5).value 
            wb1.cell(i,6).value = wb1.cell(i+1,6).value 
            wb1.cell(i,7).value = wb1.cell(i+1,7).value 
            wb1.cell(i,8).value = wb1.cell(i+1,8).value 
        wb1.cell(lastrow,2).value = None
        wb1.cell(lastrow,3).value = None
        wb1.cell(lastrow,4).value = None
        wb1.cell(lastrow,5).value = None
        wb1.cell(lastrow,6).value = None
        wb1.cell(lastrow,7).value = None
        wb1.cell(lastrow,8).value = None
        wb.save(ruta_completa +"/Registro_Imagenes.xlsx")
        print("Resultado:Se ha eliminado satisfactoriamente de GRAYSCALE")

    else:
        for i in range(3500,7000,1):
            if wb1.cell(i,8).value != None:
                lastrow = lastrow + 1
        lastrow = lastrow + 3500 - 1
        wb1.cell(row,2).value = None
        wb1.cell(row,3).value = None
        wb1.cell(row,4).value = None
        wb1.cell(row,5).value = None
        wb1.cell(row,6).value = None
        wb1.cell(row,7).value = None
        wb1.cell(row,8).value = None
        for i in range(row,lastrow,1):
            wb1.cell(i,2).value = wb1.cell(i+1,2).value
            wb1.cell(i,3).value = wb1.cell(i+1,3).value 
            wb1.cell(i,4).value = wb1.cell(i+1,4).value 
            wb1.cell(i,5).value = wb1.cell(i+1,5).value 
            wb1.cell(i,6).value = wb1.cell(i+1,6).value 
            wb1.cell(i,7).value = wb1.cell(i+1,7).value 
            wb1.cell(i,8).value = wb1.cell(i+1,8).value 
        wb1.cell(lastrow,2).value = None
        wb1.cell(lastrow,3).value = None
        wb1.cell(lastrow,4).value = None
        wb1.cell(lastrow,5).value = None
        wb1.cell(lastrow,6).value = None
        wb1.cell(lastrow,7).value = None
        wb1.cell(lastrow,8).value = None
        wb.save(ruta_completa +"/Registro_Imagenes.xlsx")
        print("Resultado: Se ha eliminado satisfactoriamente de COLOR")
    print("--------------------")

data_move2gray = ruta_completa + "/fixes/move2grayscale.csv"
data_move2color = ruta_completa + "/fixes/move2color.csv"
data_suppress = ruta_completa + "/fixes/suppress.csv"

vals_gray = pd.read_csv(data_move2gray)
vals_color = pd.read_csv(data_move2color)
vals_sup = pd.read_csv(data_suppress)

for m,val_gray in enumerate(vals_gray.values):
    move2gray(val_gray[0])
    
for n,val_color in enumerate(vals_color.values):
    move2color(val_color[0])

for o,val_sup in enumerate(vals_sup.values):
    suppress(val_sup[0])