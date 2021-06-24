from io import RawIOBase
from bs4 import BeautifulSoup
import json
import requests
import os
import pandas as pd

from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook

ruta_completa = os.getcwd()

wb = load_workbook(ruta_completa +"/Registro_Imagenes.xlsx")
wb1 = wb['Hoja1']

def scrapper(url,s):
    if s == 0:
        print("Error s = 0")
    elif s == 1:
        try:
            id_url = url.strip()[30:-1]
            val=0
            for i in range(2,7000,1):
                if wb1.cell(i,2).value == id_url:
                    val = i
            if val > 0:
                print("El url "+id_url+" ya existe")
            else: 
                print("El url "+id_url+" no existe")       
                r = requests.get(url)
                soup = BeautifulSoup(r.content,'html.parser')
                script = soup.find("script",id="initial-state")
                script_str = str(script).strip()[51:-9]
                data = json.loads(script_str)       
                row = 2
                row1 = 3502
                a=0
                for i in range(2,3500,1):
                    if wb1.cell(i,7).value != None:
                        row = i

                for i in range(0,13,1):
                    if i==0:
                        try:
                            wb1.cell(row,2).value = data['resourceResponses'][a]['response']['data']['id']
                            wb1.cell(row,3).value = data['resourceResponses'][a]['response']['data']['rich_metadata']['id']
                            wb1.cell(row,4).value = str(data['resourceResponses'][a]['response']['data']['pin_join']['visual_annotation']).strip()[1:-1]
                            wb1.cell(row,5).value = data['resourceResponses'][a]['response']['data']['origin_pinner']['full_name']
                            wb1.cell(row,6).value = data['resourceResponses'][a]['response']['data']['rich_metadata']['title']
                            wb1.cell(row,7).value = data['resourceResponses'][a]['response']['data']['rich_metadata']['description']
                            wb1.cell(row,8).value = data['resourceResponses'][a]['response']['data']['images']['orig']['url']
                            a=a + 1
                        except:
                            print("error 1") 
                            df_gray.loc[df_gray['URL']==url] = None 
                            df_gray.to_csv(ruta_completa + "/src/grayscale.csv",index=False)
                    else:
                        try:
                            wb1.cell(row,2).value = data['resourceResponses'][a]['response']['data'][i-1]['id']
                            wb1.cell(row,3).value = data['resourceResponses'][a]['response']['data'][i-1]['image_signature']
                            wb1.cell(row,4).value = str(data['resourceResponses'][a]['response']['data'][i-1]['pin_join']['visual_annotation']).strip()[1:-1]
                            wb1.cell(row,5).value = data['resourceResponses'][a]['response']['data'][i-1]['pinner']['full_name']  
                            wb1.cell(row,6).value = data['resourceResponses'][a]['response']['data'][i-1]['grid_title']
                            wb1.cell(row,7).value = data['resourceResponses'][a]['response']['data'][i-1]['description']
                            wb1.cell(row,8).value = data['resourceResponses'][a]['response']['data'][i-1]['images']['orig']['url']
                        except:
                            print("error 2")
                    row = row + 1
                    wb.save(ruta_completa +"/Registro_Imagenes.xlsx")
                    print("Se ha grabado en B/N")

        except:
            print("Error_s1")

    elif s == 2:
        try:
            id_url = url.strip()[30:-1]
            val=0
            for i in range(2,7000,1):
                if wb1.cell(i,2).value == id_url:
                    val = i
            if val > 0:
                print("El url "+id_url+" ya existe")
            else: 
                print("El url "+id_url+" no existe")       
                r = requests.get(url)
                soup = BeautifulSoup(r.content,'html.parser')
                script = soup.find("script",id="initial-state")
                script_str = str(script).strip()[51:-9]
                data = json.loads(script_str)       
                row = 2
                row1 = 3501
                a=0
                for i in range(3500,7000,1):
                    if wb1.cell(i,7).value != None:
                        row1 = i

                for i in range(0,13,1):
                    if i==0:
                        try:
                            wb1.cell(row1,2).value = data['resourceResponses'][a]['response']['data']['id']
                            wb1.cell(row1,3).value = data['resourceResponses'][a]['response']['data']['rich_metadata']['id']
                            wb1.cell(row1,4).value = str(data['resourceResponses'][a]['response']['data']['pin_join']['visual_annotation']).strip()[1:-1]
                            wb1.cell(row1,5).value = data['resourceResponses'][a]['response']['data']['origin_pinner']['full_name']   
                            wb1.cell(row1,6).value = data['resourceResponses'][a]['response']['data']['rich_metadata']['title']
                            wb1.cell(row1,7).value = data['resourceResponses'][a]['response']['data']['rich_metadata']['description']
                            wb1.cell(row1,8).value = data['resourceResponses'][a]['response']['data']['images']['orig']['url']
                            a=a + 1
                        except :
                            print("Error 3")
                            df_color.loc[df_color['URL']==url] = None
                            df_color.to_csv(ruta_completa + "/src/color.csv",index=False)

                    else:
                        try:
                            wb1.cell(row1,2).value = data['resourceResponses'][a]['response']['data'][i-1]['id']
                            wb1.cell(row1,3).value = data['resourceResponses'][a]['response']['data'][i-1]['image_signature']
                            wb1.cell(row1,4).value = str(data['resourceResponses'][a]['response']['data'][i-1]['pin_join']['visual_annotation']).strip()[1:-1]
                            wb1.cell(row1,5).value = data['resourceResponses'][a]['response']['data'][i-1]['pinner']['full_name']
                            wb1.cell(row1,6).value = data['resourceResponses'][a]['response']['data'][i-1]['grid_title']
                            wb1.cell(row1,7).value = data['resourceResponses'][a]['response']['data'][i-1]['description']
                            wb1.cell(row1,8).value = data['resourceResponses'][a]['response']['data'][i-1]['images']['orig']['url']
                        except :
                            print("Error 4")

                    row1 = row1 + 1
                    wb.save(ruta_completa +"/Registro_Imagenes.xlsx")
                    print("Se ha grabado en color")
        except:
            print("Error_s2")

    else:
        print("Error s mayor a 2")

data_gray = ruta_completa + "/src/grayscale.csv"
data_color = ruta_completa + "/src/color.csv"

urls_gray = pd.read_csv(data_gray)
urls_color = pd.read_csv(data_color)

df_gray = pd.DataFrame(urls_gray)
df_color = pd.DataFrame(urls_color)

s=0
for m,url_gray in enumerate(urls_gray.values):
    s=1
    scrapper(url_gray[0],s)
    
for n,url_color in enumerate(urls_color.values):
    s=2
    scrapper(url_color[0],s)