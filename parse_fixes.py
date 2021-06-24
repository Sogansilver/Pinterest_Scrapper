import os
import pandas as pd

from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook

ruta_completa = os.getcwd()

wb = load_workbook(ruta_completa +"/fixes/parse_url.xlsx")
wb1 = wb['Hoja1']

for i in range(1,7000,1):
    val = wb1.cell(i,1).value
    if val != None: 
        try:
            val1 = val.split('_')
            val2 = val1[1]
            wb1.cell(i,2).value = val2
            wb.save(ruta_completa +"/fixes/parse_url.xlsx")
        except:
            print("error")

print("Los URL se parsearon apropiadamente")